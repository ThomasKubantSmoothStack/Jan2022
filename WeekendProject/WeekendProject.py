from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging, os, webbrowser



logging.basicConfig(filename="log.log", filemode='w', format='%(asctime)s - %(levelname)s %(message)s', datefmt='%H:%M:%S', encoding='utf-8', level=logging.DEBUG)

months_dictionary = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12', 'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
          'May': '05', 'Jun': '06', 'Jul': '07','Aug': '08', 'Sep': '09', 'Oc': '10', 'Nov': '11', 'Dec': '12'}
years_list = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']

files = os.listdir('.')
excel_files = [item for item in files if '.xlsx' in item]

# CLI prompt that prints out list of excel files and asks user to pick one
def pick_file():
    is_picked = False
    while not is_picked:
        for item in enumerate(excel_files, 1):
            print(item)
        user_choice = input("Input number of file you would like to log: ")
        if user_choice.isdigit() and int(user_choice) <= len(excel_files):
            is_picked = True
        else:
            print("Invalid input")
    return excel_files[int(user_choice)-1][:-5]

def main():
    curr = pick_file()
    workbook = load_workbook(curr + '.xlsx')

    # date_raw extracts the month and year from the file name (for january 2018 file, date_raw is ['2018', 'january'])
    date_raw = [item for item in curr.split("_")[::-1] if item.capitalize() in months_dictionary or item in years_list]

    # date is a formatted form of date_raw (for january 2018 file, date is '2018-01')
    date = f"{date_raw[0]}-{months_dictionary[date_raw[1].capitalize()]}"
    worksheet = workbook['Summary Rolling MoM']

    # Iterate through the rows until the correct date matching the file name is found, store all the data in the row
    row = None
    for item in worksheet['A']:
        if date in str(item.value):
            row = item.row
            break
    data = [ro for ro in worksheet.iter_rows(min_row=row, max_row=row, values_only=True)]
    filtered_data = [item for item in data[0][1:] if item != None]
    logging.info(f"Month of {date_raw[1].capitalize()}, {date_raw[0]}")
    logging.info(f"Calls Offered: {filtered_data[0]}")
    logging.info(f"Abandon after 30 seconds: {round(filtered_data[1]*100,2)}%")
    logging.info(f"FCR: {filtered_data[2]*100}0%")
    logging.info(f"DSAT: {filtered_data[3]*100}0%")
    logging.info(f"CSAT: {filtered_data[4]*100}0%")

    # Switch pages, iterate through the columns until the correct date is found, store all the data in the column
    worksheet = workbook["VOC Rolling MoM"]
    col = 0
    flag = False
    try:
        for item in worksheet['1']:
            col += 1
            if date in str(item.value) or str(item.value) in months_dictionary:
                flag = True
                break
    except Exception as e:
        logging.warning(e)
    finally:
        if not flag:
            logging.warning("Unable to find correct date column. Please ensure every column matches standard naming.")
    scores = [item.value for item in worksheet[get_column_letter(col)][1:]]
    print(scores)
    if scores[2] > 200:
        logging.info(f"Promoters: {scores[2]}: GOOD")
    else:
        logging.info(f"Promoters: {scores[2]}: BAD")
    if scores[4] > 100:
        logging.info(f"Passives: {scores[4]}: GOOD")
    else:
        logging.info(f"Passives: {scores[4]}: BAD")
    if scores[6] > 100:
        logging.info(f"Detractors: {scores[6]}: GOOD")
    else:
        logging.info(f"Detractors: {scores[6]}: BAD")
    webbrowser.open("log.log")

if __name__ == "__main__":
    main()